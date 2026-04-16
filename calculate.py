import pandas as pd
import io
from db import students, uploads
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import copy


# ─── helpers ─────────────────────────────────────────────────────────────────

def _thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=11):
    return Font(name="Arial", bold=bold, size=size)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _set(ws, row, col, value, bold=False, size=11, h="center", v="center",
         fill=None, wrap=False, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font  = _font(bold, size)
    c.alignment = _align(h, v, wrap)
    if fill:
        c.fill = _fill(fill)
    if border:
        c.border = _thin()
    return c

def _merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2, end_column=c2)


# ─── main ────────────────────────────────────────────────────────────────────

def calculate_attainment(threshold, max_marks,
                          co_maxes=None,
                          teacher=None,
                          session_students=None,
                          session_uploads=None,
                          program="", semester="", course_code="", course_name="",
                          department="Department of Management Studies",
                          university="CT UNIVERSITY",
                          exam_label="Mid Term Examination"):
    """
    Build a CO Attainment Excel that mirrors the template layout.
    co_maxes: list of max marks per CO (e.g. [5, 4, 6]). If None, splits evenly.
    Returns (df, summary).
    """

    # ── fetch data: use session data if provided, otherwise query MongoDB ──────
    if session_students is not None or session_uploads is not None:
        # Only use data from the current session
        raw_data = list(session_students or []) + list(session_uploads or [])
    else:
        # Fallback: query MongoDB filtered by teacher
        q = {"teacher": teacher} if teacher else {}
        raw_data = list(students.find(q)) + list(uploads.find(q))

    def normalise(s):
        name = (s.get("student") or s.get("Student Name") or
                s.get("Name") or s.get("student_name") or "")
        reg_no = (s.get("reg_no") or s.get("Reg No") or
                  s.get("RegNo") or s.get("Registration No") or "")
        status = (s.get("status") or s.get("Attendance") or "Present")
        if "co_marks" in s:
            co_vals = [float(v or 0) for v in s["co_marks"]]
        else:
            co_vals, i = [], 1
            while True:
                v = (s.get(f"CO{i}") if s.get(f"CO{i}") is not None else
                     s.get(f"CO{i} Marks") if s.get(f"CO{i} Marks") is not None else
                     s.get(f"co{i}") if s.get(f"co{i}") is not None else None)
                if v is None:
                    break
                co_vals.append(float(v or 0))
                i += 1
        return {"name": name, "reg_no": reg_no, "status": status, "co_vals": co_vals}

    data = [normalise(s) for s in raw_data]

    # ── detect number of COs ──────────────────────────────────────────────────
    num_co = max((len(d["co_vals"]) for d in data), default=3)
    num_co = max(num_co, 3)

    # ── max-marks per CO ──────────────────────────────────────────────────────
    if co_maxes and len(co_maxes) >= num_co:
        # Use exactly what the teacher set — validated by UI co_sum check
        co_maxes = [float(v) for v in co_maxes[:num_co]]
    else:
        # Fallback: split evenly if no co_maxes provided
        base = round(max_marks / num_co, 1)
        co_maxes = [base] * num_co

    # ── build workbook ────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "First Method"

    # column widths  (A…K equivalent, extended for extra COs)
    # A,B,C/D,E,F,G,H,I,J,K …
    base_widths = [6, 14, 20, 8, 10, 10, 10, 10, 10, 10, 10]
    extra = max(0, num_co - 3)
    col_widths = base_widths + [10] * extra
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # total columns needed: A(1) B(2) C+D(3-4) E(5)  F..F+num_co-1 (CO raw)
    #                        F+num_co .. F+2*num_co-1 (CO %)
    # Let's use a fixed layout matching the template exactly (extended for COs)
    # Columns: 1=Sr, 2=Reg, 3-4=Name(merged), 5=Total, 6..5+num_co=CO raw, 6+num_co..5+2*num_co=CO%
    COL_SR    = 1
    COL_REG   = 2
    COL_NAME  = 3   # merged with 4
    COL_TOTAL = 5
    COL_CO1_RAW = 6
    COL_CO1_PCT = 6 + num_co
    LAST_COL    = 5 + 2 * num_co

    # ── ROW 1: University ─────────────────────────────────────────────────────
    _merge(ws, 1, 1, 1, LAST_COL)
    _set(ws, 1, 1, university, bold=True, size=12)

    # ── ROW 2: Department ─────────────────────────────────────────────────────
    _merge(ws, 2, 1, 2, LAST_COL)
    _set(ws, 2, 1, department, bold=True, size=12)

    # ── ROWS 3-5: Attainment criteria + course info ───────────────────────────
    yellow = "FFFF00"
    pink   = "FEB4B4"

    # criteria block (col A-B merged, rows 3-5)
    _merge(ws, 3, 1, 5, 2)
    _set(ws, 3, 1, "ATTAINMENT CRITERIA", size=10, h="center", v="center")

    for row, pct, level, label in [(3, 70, 3, "Very Good"),
                                    (4, 60, 2, "Good"),
                                    (5, 50, 1, "Average")]:
        _set(ws, row, 3, pct,   bold=True,  size=10, fill=yellow if row==4 else None)
        _set(ws, row, 4, level, size=10,    fill=yellow if row==4 else None)
        _set(ws, row, 5, label, size=10,    fill=yellow if row==4 else None)

    # threshold % label + value
    _merge(ws, 3, 6, 3, 8)
    _set(ws, 3, 6, "Threshold % for attainment", size=10)
    _merge(ws, 3, 9, 3, LAST_COL)
    _set(ws, 3, 9, threshold, bold=True, size=10, fill=yellow)

    # Program / Sem
    _merge(ws, 4, 6, 4, 8)
    _set(ws, 4, 6, "Program", bold=True, size=10)
    _merge(ws, 4, 9, 4, 10)
    _set(ws, 4, 9, program or "BBA", bold=True, size=10, fill=yellow)
    _set(ws, 4, 11, "Sem.", bold=True, size=10)
    _merge(ws, 4, 12, 4, LAST_COL) if LAST_COL >= 12 else None
    _set(ws, 4, 12 if LAST_COL >= 12 else LAST_COL, semester or "II", bold=True, size=10, fill=yellow)

    # Course Code / Name
    _merge(ws, 5, 6, 5, 8)
    _set(ws, 5, 6, "Course Code", bold=True, size=10)
    _merge(ws, 5, 9, 5, 10)
    _set(ws, 5, 9, course_code or "", size=10)
    _set(ws, 5, 11, "Course Name", bold=True, size=10)
    _merge(ws, 5, 12, 5, LAST_COL) if LAST_COL >= 12 else None
    _set(ws, 5, 12 if LAST_COL >= 12 else LAST_COL, course_name or "", bold=True, size=10)

    # ── ROW 6: Section header ──────────────────────────────────────────────────
    _merge(ws, 6, 1, 6, LAST_COL)
    _set(ws, 6, 1, f"Attainment for {exam_label} (Sessional)", bold=True, size=11)

    # ── ROWS 7-8: Column headers ───────────────────────────────────────────────
    _merge(ws, 7, 1, 8, 4)
    _set(ws, 7, 1, "CIA", bold=True, size=11)

    _merge(ws, 7, 5, 8, 5)
    _set(ws, 7, 5, "Max marks", bold=True, size=11)

    # raw CO headers (row 7 label, row 8 CO names)
    _merge(ws, 7, COL_CO1_RAW, 7, COL_CO1_RAW + num_co - 1)
    _set(ws, 7, COL_CO1_RAW, exam_label, bold=True, size=11)
    for i in range(num_co):
        _set(ws, 8, COL_CO1_RAW + i, f"CO{i+1}", bold=True, size=11)

    # % CO headers
    for i in range(num_co):
        _set(ws, 7, COL_CO1_PCT + i, f"CO{i+1}", bold=True, size=11)
        _set(ws, 8, COL_CO1_PCT + i, "%",         bold=True, size=11)

    # ── ROW 9: Max-marks row ───────────────────────────────────────────────────
    _merge(ws, 9, 1, 9, 4)
    _set(ws, 9, 1, "Sr.No / Reg / Name", size=11)
    _set(ws, 9, 5, max_marks, bold=True, size=11)
    for i, mx in enumerate(co_maxes):
        _set(ws, 9, COL_CO1_RAW + i, mx,  bold=True, size=11)
        _set(ws, 9, COL_CO1_PCT + i, 100, bold=True, size=11)

    # ── ROWS 10+: Student data ─────────────────────────────────────────────────
    start_row  = 10
    rows_out   = []
    passed_co  = [0] * num_co
    absent_co  = [0] * num_co
    total_present = 0

    for idx, s in enumerate(data):
        r = start_row + idx

        name   = s["name"]
        reg_no = s["reg_no"]
        status = s["status"]
        co_vals = [s["co_vals"][i] if i < len(s["co_vals"]) else 0.0
                   for i in range(num_co)]

        total = sum(co_vals)

        # Sr, Reg, Name
        _set(ws, r, COL_SR,  idx + 1)
        _set(ws, r, COL_REG, reg_no)
        _merge(ws, r, COL_NAME, r, 4)
        _set(ws, r, COL_NAME, name)
        _set(ws, r, COL_TOTAL, total)

        # CO raw (write actual values; formulas not needed since we know the marks)
        for i, v in enumerate(co_vals):
            _set(ws, r, COL_CO1_RAW + i, v)

        # CO % — write actual calculated values (not formulas) so Excel shows
        # correct numbers without needing recalculation
        co_pcts = []
        for i in range(num_co):
            if status == "Absent":
                _set(ws, r, COL_CO1_PCT + i, "AB", fill="FEB4B4")
                absent_co[i] += 1
                co_pcts.append(0)
            else:
                pct = min(round(co_vals[i] / co_maxes[i] * 100, 2), 100.0) if co_maxes[i] > 0 else 0
                _set(ws, r, COL_CO1_PCT + i, pct)
                co_pcts.append(pct)
                if pct >= threshold:
                    passed_co[i] += 1

        if status != "Absent":
            total_present += 1

        total_pct = (total / max_marks * 100) if max_marks > 0 else 0
        rows_out.append({
            "Student":     name,
            "Total Marks": total,
            "Total %":     round(total_pct, 2),
            "Status":      "Absent" if status == "Absent" else
                           ("Pass" if total_pct >= threshold else "Fail")
        })

    last_data_row = start_row + len(data) - 1
    pct_col = lambda i: f"${get_column_letter(COL_CO1_PCT + i)}"

    # ── ATTAINMENT TABLE rows ──────────────────────────────────────────────────
    tbl_row = last_data_row + 1

    # ── Pre-calculate all summary values from rows_out & passed_co ─────────────
    present_rows = [d for d in data if d["status"] != "Absent"]
    num_present  = len(present_rows)

    # Per-CO: average % across present students
    co_avg_pcts = []
    for i in range(num_co):
        vals_pct = []
        for d in present_rows:
            v = d["co_vals"][i] if i < len(d["co_vals"]) else 0
            vals_pct.append(min(round(v / co_maxes[i] * 100, 2), 100.0) if co_maxes[i] > 0 else 0)
        co_avg_pcts.append(round(sum(vals_pct) / len(vals_pct), 2) if vals_pct else 0)

    # Per-CO: count present students scoring >= threshold
    co_above_thresh = []
    for i in range(num_co):
        cnt = 0
        for d in present_rows:
            v = d["co_vals"][i] if i < len(d["co_vals"]) else 0
            pct = min(round(v / co_maxes[i] * 100, 2), 100.0) if co_maxes[i] > 0 else 0
            if pct >= threshold:
                cnt += 1
        co_above_thresh.append(cnt)

    # Per-CO: % of present students above threshold
    co_pct_above = [
        round(co_above_thresh[i] / num_present * 100, 2) if num_present > 0 else 0
        for i in range(num_co)
    ]

    # Per-CO: attainment level 1/2/3
    co_attainment = [
        3 if co_pct_above[i] >= 70 else (2 if co_pct_above[i] >= 60 else (1 if co_pct_above[i] >= 50 else 0))
        for i in range(num_co)
    ]

    # Absent count per CO
    absent_counts = [sum(1 for d in data if d["status"] == "Absent") for _ in range(num_co)]

    # Sessional attainment = average of per-CO attainment levels
    sessional_attainment = round(sum(co_attainment) / num_co, 2) if num_co > 0 else 0

    # Row A: Attainment through Sessional
    _merge(ws, tbl_row, 1, tbl_row, 4)
    _set(ws, tbl_row, 1, "Attainment through Sessional Examination:", bold=True, size=9, fill="D9EAD3")
    _set(ws, tbl_row, 5, sessional_attainment, bold=True, size=9, fill="D9EAD3")

    _merge(ws, tbl_row, 6, tbl_row, 6 + num_co - 1) if num_co > 1 else None
    _set(ws, tbl_row, 6, "ATTAINMENT TABLE", size=9, fill=yellow)

    for i in range(num_co):
        _set(ws, tbl_row, COL_CO1_PCT + i, absent_counts[i], size=9, fill=yellow)

    # Row B: Present students
    tbl_row += 1
    _merge(ws, tbl_row, 1, tbl_row, 4)
    _set(ws, tbl_row, 1, "Attainment through university examination:", bold=True, size=9)
    _set(ws, tbl_row, 5, "NA", bold=True, size=9)
    _merge(ws, tbl_row, 6, tbl_row, 6 + num_co - 1) if num_co > 1 else None
    _set(ws, tbl_row, 6, "PRESENT STUDENT OR ATTEMPT", size=8, fill="D9D9D9")
    for i in range(num_co):
        _set(ws, tbl_row, COL_CO1_PCT + i, num_present, size=9, fill=yellow)

    # Row C: No. of students above threshold
    tbl_row += 1
    _merge(ws, tbl_row, 1, tbl_row, 4)
    _set(ws, tbl_row, 1, "Weightage given to the Internal examination (40%):", size=9, fill=pink)
    _set(ws, tbl_row, 5, "NA", bold=True, size=9, fill=pink)
    _merge(ws, tbl_row, 6, tbl_row, 6 + num_co - 1) if num_co > 1 else None
    _set(ws, tbl_row, 6, "NO. OF STUDENTS SECURE MARKS > THRESHOLD MARKS", size=8, fill="D9D9D9")
    for i in range(num_co):
        _set(ws, tbl_row, COL_CO1_PCT + i, co_above_thresh[i], size=9, fill=yellow)

    # Row D: % of students above threshold
    tbl_row += 1
    _merge(ws, tbl_row, 1, tbl_row, 4)
    _set(ws, tbl_row, 1, "Weightage given to the university examination (60%):", size=9, fill=pink)
    _set(ws, tbl_row, 5, "NA", bold=True, size=9, fill=pink)
    _merge(ws, tbl_row, 6, tbl_row, 6 + num_co - 1) if num_co > 1 else None
    _set(ws, tbl_row, 6, "% OF STUDENTS SECURE MARKS > THRESHOLD MARKS", size=8, fill="D9D9D9")
    for i in range(num_co):
        _set(ws, tbl_row, COL_CO1_PCT + i, co_pct_above[i], size=9, fill=yellow)

    # Row E: Final attainment level
    tbl_row += 1
    _merge(ws, tbl_row, 1, tbl_row, 4)
    _set(ws, tbl_row, 1, "Final attainment level of the course (by Direct Assessment):",
         size=9, h="left", fill="D9D9D9")
    _set(ws, tbl_row, 5, "NA", bold=True, size=9)
    _merge(ws, tbl_row, 6, tbl_row, 6 + num_co - 1) if num_co > 1 else None
    _set(ws, tbl_row, 6, "Attainment (3 ≥ 70%, 2 ≥ 60%, 1 ≥ 50%)", size=8, fill="D9D9D9")
    for i in range(num_co):
        _set(ws, tbl_row, COL_CO1_PCT + i, co_attainment[i], size=9, fill=yellow)

    # ── summary ───────────────────────────────────────────────────────────────
    total_students = len(rows_out)
    passed_total   = sum(1 for r in rows_out if r["Status"] == "Pass")
    attainment_pct = round(passed_total / total_students * 100, 2) if total_students else 0

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    df = pd.DataFrame(rows_out)
    summary = pd.DataFrame({
        "Parameter": ["Total Students", "Passed", "Attainment %"],
        "Value":     [total_students, passed_total, attainment_pct]
    })
    return df, summary, excel_buffer