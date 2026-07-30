import streamlit as st
import matplotlib.pyplot as plt
import pandas as pd
from db import students, users, uploads
from calculate import calculate_attainment
import os
import time 
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


# ---------------- TEMPLATE BUILDER ----------------
def _build_template():
    """Generate the student data entry template as a BytesIO object."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection

    NUM_STUDENTS = 30
    BLUE_HDR = "1E3A5F"
    ALT1     = "EBF5FB"
    ALT2     = "FFFFFF"

    def S(): return Side(style="thin")
    def bdr(): return Border(left=S(), right=S(), top=S(), bottom=S())

    def c(ws, r, col, value=None, bold=False, size=11, h="center", v="center",
          bg=None, locked=True, color="000000"):
        cl = ws.cell(row=r, column=col, value=value)
        cl.font       = Font(name="Arial", bold=bold, size=size, color=color)
        cl.alignment  = Alignment(horizontal=h, vertical=v)
        if bg: cl.fill = PatternFill("solid", fgColor=bg)
        cl.border     = bdr()
        cl.protection = Protection(locked=locked)
        return cl

    def m(ws, r1, c1, r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    wb = Workbook()
    ws = wb.active
    ws.title = "Student Data"

    for col, w in zip("ABCDEFG", [6, 14, 26, 8, 12, 12, 12]):
        ws.column_dimensions[col].width = w

    # Row 1: Headers
    ws.row_dimensions[1].height = 22
    for c1, c2, label in [
        (1,1,"Sr.No"),(2,2,"Reg No"),(3,4,"Student Name"),
        (5,5,"CO1 Marks"),(6,6,"CO2 Marks"),(7,7,"CO3 Marks")
    ]:
        if c1 != c2: m(ws, 1, c1, 1, c2)
        c(ws, 1, c1, label, bold=True, size=11, bg=BLUE_HDR, color="FFFFFF")

    # Student rows
    for idx in range(NUM_STUDENTS):
        r   = 2 + idx
        alt = ALT1 if idx % 2 == 0 else ALT2
        ws.row_dimensions[r].height = 17
        c(ws, r, 1, idx+1, size=10, bg=alt)
        c(ws, r, 2, None,  size=10, bg=ALT2, locked=False, h="center")
        m(ws, r, 3, r, 4)
        c(ws, r, 3, None,  size=10, bg=ALT2, locked=False, h="left")
        c(ws, r, 5, None,  size=10, bg=ALT2, locked=False)
        c(ws, r, 6, None,  size=10, bg=ALT2, locked=False)
        c(ws, r, 7, None,  size=10, bg=ALT2, locked=False)

    ws.freeze_panes = "A2"
    # No sheet protection — teacher can freely edit all cells

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------- LOAD CUSTOM CSS ----------------
def load_css():
    """Load external CSS file and hide default Streamlit UI elements"""
    with open("styles.css") as f:
        st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)

    st.markdown("""
    <style>
    #MainMenu {visibility: hidden;}
    header {visibility: hidden;}
    </style>
    """, unsafe_allow_html=True)

load_css()


# ---------------- FOOTER ----------------
def show_footer():
    st.markdown("""
        <div class="custom-footer">
            Developed by Siddharth Singh Bhandari | Attainment System 🚀
        </div>
    """, unsafe_allow_html=True)


# ---------------- SESSION STATE ----------------
if "page" not in st.session_state:
    st.session_state.page = "home"

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

# ✅ ADD (for reset)
if "reset_key" not in st.session_state:
    st.session_state.reset_key = 0

# ✅ ADD (flicker fix)
if "just_logged_in" not in st.session_state:
    st.session_state.just_logged_in = False


# ---------------- NAVIGATION BAR ----------------
def navbar():
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        if st.button("Admin Register", use_container_width=True):
            st.session_state.page = "admin_register"

    with col2:
        if st.button("Admin Login", use_container_width=True):
            st.session_state.page = "admin_login"

    with col3:
        if st.button("Teacher Register", use_container_width=True):
            st.session_state.page = "teacher_register"

    with col4:
        if st.button("Teacher Login", use_container_width=True):
            st.session_state.page = "teacher_login"

navbar()

# ✅ FIX (no stop → no blank page)
if st.session_state.just_logged_in:
    st.session_state.just_logged_in = False


# ---------------- APP TITLE ----------------
st.title("Attainment System", anchor=False)


# ---------------- LOGO DISPLAY ----------------
logo_url = "https://fimt-ggsipu.org/images/flogo2025-1.jpg"
col1, col2, col3 = st.columns([1, 4, 1])
with col2:
    st.image(logo_url, width=450)


# ---------------- ADMIN REGISTER ----------------
# ─────────────────────────────────────────────────────────
# ADMIN REGISTER
# ─────────────────────────────────────────────────────────
if st.session_state.page == "admin_register":

    _, col, _ = st.columns([1, 2, 1])
    with col:
        with st.container(border=True):
            st.markdown("### 🛡️ Create Admin Account")
            st.markdown("Register a new administrator for this system.")
            st.markdown("---")
            with st.form("admin_register_form"):
                username = st.text_input("👤 Admin Username", placeholder="Enter username")
                password = st.text_input("🔒 Password", type="password", placeholder="Enter password")
                confirm  = st.text_input("🔒 Confirm Password", type="password", placeholder="Re-enter password")
                submitted = st.form_submit_button("Create Admin Account", use_container_width=True)

                if submitted:
                    if not username.strip() or not password.strip():
                        st.error("❌ All fields are required.")
                    elif password != confirm:
                        st.error("❌ Passwords do not match.")
                    elif users.find_one({"username": username.strip(), "role": "admin"}):
                        st.error("❌ Admin with this username already exists.")
                    else:
                        users.insert_one({"username": username.strip(), "password": password, "role": "admin"})
                        st.success("✅ Admin account created successfully!")

        st.markdown("")
        if st.button("← Back to Home", use_container_width=True):
            st.session_state.page = "home"
            st.rerun()


# ─────────────────────────────────────────────────────────
# ADMIN LOGIN
# ─────────────────────────────────────────────────────────
elif st.session_state.page == "admin_login":

    _, col, _ = st.columns([1, 2, 1])
    with col:
        with st.container(border=True):
            st.markdown("### 🛡️ Admin Login")
            st.markdown("Sign in to manage teachers and system settings.")
            st.markdown("---")
            with st.form("admin_login_form"):
                username  = st.text_input("👤 Username", placeholder="Enter your username")
                password  = st.text_input("🔒 Password", type="password", placeholder="Enter your password")
                submitted = st.form_submit_button("Login as Admin", use_container_width=True)

                if submitted:
                    if not username.strip() or not password.strip():
                        st.error("❌ Please fill in all fields.")
                    else:
                        user = users.find_one({"username": username.strip(), "password": password, "role": "admin"})
                        if user:
                            st.session_state.logged_in      = True
                            st.session_state.admin_username = username.strip()
                            st.session_state.page           = "admin_panel"
                            st.session_state.just_logged_in = True
                            st.rerun()
                        else:
                            st.error("❌ Invalid username or password.")

        st.markdown("")
        if st.button("← Back to Home", use_container_width=True):
            st.session_state.page = "home"
            st.rerun()


# ─────────────────────────────────────────────────────────
# ADMIN DASHBOARD
# ─────────────────────────────────────────────────────────
elif st.session_state.page == "admin_panel" and st.session_state.logged_in:

    admin_name = st.session_state.get("admin_username", "Admin")
    st.header(f"🛡️ Admin Dashboard", anchor=False)
    st.markdown(f"Welcome, **{admin_name}**")
    st.markdown("---")

    # ── Stats row ──────────────────────────────────────────
    teachers_list = list(users.find({"role": "teacher"}))
    admins_list   = list(users.find({"role": "admin"}))

    c1, c2, c3 = st.columns(3)
    c1.metric("👨‍🏫 Teachers",  len(teachers_list))
    c2.metric("🛡️ Admins",    len(admins_list))
    c3.metric("👥 Total Users", len(teachers_list) + len(admins_list))

    st.markdown("")

    # ── Teachers ───────────────────────────────────────────
    with st.container(border=True):
        st.subheader("👨‍🏫 Registered Teachers", anchor=False)
        if teachers_list:
            for t in teachers_list:
                col1, col2 = st.columns([5, 1])
                with col1:
                    st.markdown(f"**{t['username']}**")
                with col2:
                    if st.button("🗑️ Delete", key=f"del_t_{t['_id']}"):
                        users.delete_one({"_id": t["_id"]})
                        st.success(f"Teacher '{t['username']}' removed.")
                        st.rerun()
        else:
            st.info("No teachers registered yet.")

    # ── Admins ─────────────────────────────────────────────
    with st.container(border=True):
        st.subheader("🛡️ Registered Admins", anchor=False)
        for a in admins_list:
            col1, col2 = st.columns([5, 1])
            with col1:
                badge = " *(you)*" if a["username"] == admin_name else ""
                st.markdown(f"**{a['username']}**{badge}")
            with col2:
                if a["username"] != admin_name:
                    if st.button("🗑️ Delete", key=f"del_a_{a['_id']}"):
                        users.delete_one({"_id": a["_id"]})
                        st.success(f"Admin '{a['username']}' removed.")
                        st.rerun()
                else:
                    st.caption("(current)")

    st.markdown("---")
    if st.button("🚪 Logout", use_container_width=True):
        st.session_state.logged_in = False
        st.session_state.page      = "home"
        st.rerun()


# ─────────────────────────────────────────────────────────
# TEACHER REGISTER
# ─────────────────────────────────────────────────────────
elif st.session_state.page == "teacher_register":

    _, col, _ = st.columns([1, 2, 1])
    with col:
        with st.container(border=True):
            st.markdown("### 👨‍🏫 Create Teacher Account")
            st.markdown("Register a new teacher to use the attainment system.")
            st.markdown("---")
            with st.form("teacher_register_form"):
                username  = st.text_input("👤 Username", placeholder="Choose a username")
                password  = st.text_input("🔒 Password", type="password", placeholder="Create a password")
                confirm   = st.text_input("🔒 Confirm Password", type="password", placeholder="Re-enter password")
                submitted = st.form_submit_button("Register Teacher", use_container_width=True)

                if submitted:
                    if not username.strip() or not password.strip():
                        st.error("❌ All fields are required.")
                    elif password != confirm:
                        st.error("❌ Passwords do not match.")
                    elif users.find_one({"username": username.strip()}):
                        st.error("❌ Username already taken. Choose another.")
                    else:
                        users.insert_one({"username": username.strip(), "password": password, "role": "teacher"})
                        st.success("✅ Teacher account created! You can now log in.")

        st.markdown("")
        if st.button("← Back to Home", use_container_width=True):
            st.session_state.page = "home"
            st.rerun()


# ─────────────────────────────────────────────────────────
# TEACHER LOGIN
# ─────────────────────────────────────────────────────────
elif st.session_state.page == "teacher_login":

    _, col, _ = st.columns([1, 2, 1])
    with col:
        with st.container(border=True):
            st.markdown("### 👨‍🏫 Teacher Login")
            st.markdown("Sign in to manage students and calculate attainment.")
            st.markdown("---")
            with st.form("teacher_login_form"):
                username  = st.text_input("👤 Username", placeholder="Enter your username")
                password  = st.text_input("🔒 Password", type="password", placeholder="Enter your password")
                submitted = st.form_submit_button("Login as Teacher", use_container_width=True)

                if submitted:
                    if not username.strip() or not password.strip():
                        st.error("❌ Please fill in all fields.")
                    else:
                        user = users.find_one({"username": username.strip(), "password": password, "role": "teacher"})
                        if user:
                            st.session_state.logged_in      = True
                            st.session_state.teacher        = username.strip()
                            st.session_state.page           = "teacher_panel"
                            st.session_state.just_logged_in = True
                            st.session_state["session_upload_df"]       = None
                            st.session_state["session_manual_students"] = []
                            st.session_state["_session_students"]       = []
                            st.session_state["_session_uploads"]        = []
                            st.rerun()
                        else:
                            st.error("❌ Invalid username or password.")

        st.markdown("")
        if st.button("← Back to Home", use_container_width=True):
            st.session_state.page = "home"
            st.rerun()


# ---------------- TEACHER DASHBOARD ----------------
elif st.session_state.page == "teacher_panel" and st.session_state.logged_in:

    st.header("Teacher Dashboard", anchor=False)
    st.markdown("---")

    # ── BOX 1: TEMPLATE & FILE UPLOAD ──────────────────
    with st.container(border=True):
        st.subheader("📂 Template & Upload", anchor=False)
        col_dl, col_up = st.columns([1, 2])
        with col_dl:
            st.markdown("**Step 1 — Download the template:**")
            st.download_button(
                "⬇️ Download Student Data Template",
                data=_build_template(),
                file_name="Student_Data_Template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        with col_up:
            st.markdown("**Step 2 — Upload filled file:**")
            uploaded_file = st.file_uploader("Upload Excel/CSV", type=["xlsx", "csv"], label_visibility="collapsed")

        if uploaded_file is not None:
            if uploaded_file.name.endswith(".csv"):
                df_upload = pd.read_csv(uploaded_file)
            else:
                df_upload = pd.read_excel(uploaded_file, engine="openpyxl")
            st.dataframe(df_upload, use_container_width=True)

            if "data_saved" not in st.session_state:
                st.session_state.data_saved = False

            if st.button("💾 Save Uploaded Data", use_container_width=True):
                # Drop rows with no student name (blank template rows)
                df_upload = df_upload.dropna(subset=[c for c in df_upload.columns if "name" in c.lower() or "student" in c.lower()], how="all")
                df_upload = df_upload[df_upload.apply(lambda r: any(str(v).strip() not in ("","nan","None") for v in r), axis=1)]
                rows_to_save = []
                for row in df_upload.to_dict(orient="records"):
                    row["teacher"] = st.session_state.teacher
                    row["status"]  = str(row["Attendance"]).strip() if "Attendance" in row else "Present"
                    if "Max Marks" in row:
                        try:    row["Max Marks"] = float(row["Max Marks"])
                        except: row["Max Marks"] = 15
                    uploads.insert_one(row)
                    rows_to_save.append(row)
                # Store this session's uploaded data so download only shows THIS upload
                st.session_state["session_upload_df"] = df_upload.copy()
                st.session_state.data_saved = True

            if st.session_state.data_saved:
                st.success("✅ Data saved successfully")
                st.session_state.data_saved = False

        # Clear session data button
        has_session = (
            st.session_state.get("session_upload_df") is not None or
            bool(st.session_state.get("session_manual_students"))
        )
        if has_session:
            if st.button("🗑️ Clear Session Data (start fresh)", use_container_width=True):
                st.session_state["session_upload_df"]       = None
                st.session_state["session_manual_students"] = []
                st.session_state["_session_students"]       = []
                st.session_state["_session_uploads"]        = []
                st.success("Session cleared. Upload new data to start fresh.")
                st.rerun()

    # ── BOX 2: CLASS CONFIGURATION ─────────────────────
    with st.container(border=True):
        st.subheader("📊 Class Configuration", anchor=False)
        col_a, col_b = st.columns(2)
        with col_a:
            total_students_input = st.number_input("Total Students", min_value=1, value=1)
            max_marks_input      = st.number_input("Maximum Marks (Total)", min_value=1, value=15)
        with col_b:
            threshold_global = st.number_input(
                "Threshold %", min_value=0, max_value=100, value=50,
                help="Students scoring >= this % per CO are counted as attained"
            )
            co_no_config = int(st.number_input("Number of COs", min_value=1, max_value=6, value=3, step=1))

        st.markdown("**Set max marks per CO** (must sum to Total Max Marks)")
        co_max_cols    = st.columns(co_no_config)
        co_maxes_input = []
        for i in range(co_no_config):
            with co_max_cols[i]:
                val = st.number_input(
                    f"CO{i+1} Max",
                    min_value=0.0, max_value=float(max_marks_input),
                    value=float(round(max_marks_input / co_no_config, 1)),
                    step=0.5, key=f"co_max_{i}"
                )
                co_maxes_input.append(val)

        co_sum = sum(co_maxes_input)
        if abs(co_sum - max_marks_input) > 0.01:
            st.warning(f"CO sum = {co_sum} does not match Total Max = {max_marks_input}")
        else:
            st.success(f"✅ {' + '.join(str(v) for v in co_maxes_input)} = {co_sum}")

    # ── BOX 3: COURSE & INSTITUTE DETAILS ──────────────
    with st.container(border=True):
        st.subheader("🏫 Course & Institute Details", anchor=False)
        col_a, col_b = st.columns(2)
        with col_a:
            university_name = st.text_input("University Name", value="CT UNIVERSITY")
            department_name = st.text_input("Department",      value="Department of Management Studies")
            program_name    = st.text_input("Program",         value="BBA")
        with col_b:
            semester_val    = st.text_input("Semester",        value="II")
            course_code_val = st.text_input("Course Code")
            course_name_val = st.text_input("Course Name")
        exam_label_val = st.text_input("Exam Label", value="Mid Term Examination")

    # ── BOX 4: MANUAL STUDENT ENTRY ────────────────────
    with st.container(border=True):
        st.subheader("✏️ Manual Student Entry", anchor=False)

        co_no = st.number_input("Number of CO", min_value=0, max_value=6, step=1)
        co_marks_inputs = []
        for i in range(int(co_no)):
            co_marks_inputs.append(
                st.text_input(f"CO{i+1} Marks", key=f"co_{i}_{st.session_state.reset_key}")
            )

        with st.form(f"student_form_{st.session_state.reset_key}"):
            subject   = st.text_input("Subject Name")
            code      = st.text_input("Subject Code")
            threshold = st.number_input("Threshold Marks", min_value=0, value=0)
            name      = st.text_input("Student Name")
            status    = st.selectbox("Attendance", ["Present", "Absent"])
            submit_student = st.form_submit_button("➕ Add Student", use_container_width=True)

            if submit_student:
                if name.strip() == "" or subject.strip() == "" or code.strip() == "":
                    st.error("❌ Student Name, Subject Name and Subject Code are required")
                else:
                    marks_int = [int(x) if x.isdigit() else 0 for x in co_marks_inputs]
                    total     = sum(marks_int)
                    new_student = {
                        "student": name, "teacher": st.session_state.teacher,
                        "subject": subject, "code": code,
                        "threshold": int(threshold), "co_marks": marks_int,
                        "total": total, "final_total": total, "status": status
                    }
                    students.insert_one(new_student)
                    # Track manually added students for this session
                    if "session_manual_students" not in st.session_state:
                        st.session_state["session_manual_students"] = []
                    st.session_state["session_manual_students"].append(new_student)
                    st.success("✅ Student Added Successfully")
                    time.sleep(1)
                    st.session_state.reset_key += 1
                    st.rerun()

    # ── BOX 5: DOWNLOAD & CALCULATE ATTAINMENT ─────────
    with st.container(border=True):
        st.subheader("📥 Download & Calculate Attainment", anchor=False)

        col1, col2 = st.columns(2)
        with col1:
            download_clicked  = st.button("📊 Download My Students Data", use_container_width=True)
        with col2:
            calculate_clicked = st.button("🎯 Calculate Attainment",       use_container_width=True)

        if download_clicked:
            # Use only data from THIS session (uploaded file OR manually added)
            # Not all historical data from MongoDB
            session_manual = st.session_state.get("session_manual_students", [])
            session_upload_df = st.session_state.get("session_upload_df", None)

            all_data = list(session_manual)
            if session_upload_df is not None:
                # Filter out rows with no student name or reg no (empty template rows)
                clean_df = session_upload_df[
                    session_upload_df.apply(
                        lambda r: str(r.get("Student Name", r.get("student_name", r.get("Name", "")))).strip().lower()
                                  not in ("", "nan", "none"),
                        axis=1
                    )
                ]
                for row in clean_df.to_dict(orient="records"):
                    row["teacher"] = st.session_state.get("teacher", "")
                    row["status"]  = str(row.get("Attendance", "Present")).strip()
                    all_data.append(row)

            if not all_data:
                st.warning("⚠️ No data for this session. Upload a file or add students manually first.")
            else:
                def normalise(d):
                    name   = (d.get("student") or d.get("Student Name") or d.get("Name") or d.get("student_name") or "")
                    reg_no = (d.get("reg_no") or d.get("Reg No") or d.get("RegNo") or d.get("Registration No") or "")
                    status = (d.get("status") or d.get("Attendance") or "Present")
                    if "co_marks" in d:
                        co_vals = [float(v or 0) for v in d["co_marks"]]
                    else:
                        co_vals, i = [], 1
                        while True:
                            v = (d.get(f"CO{i}") if d.get(f"CO{i}") is not None else
                                 d.get(f"CO{i} Marks") if d.get(f"CO{i} Marks") is not None else
                                 d.get(f"co{i}") if d.get(f"co{i}") is not None else None)
                            if v is None: break
                            co_vals.append(float(v or 0))
                            i += 1
                    total = d.get("total") if d.get("total") else sum(float(v or 0) for v in co_vals)
                    return {
                        "name": name, "reg_no": reg_no, "status": status,
                        "subject": d.get("subject") or d.get("Subject Name", ""),
                        "code":    d.get("code")    or d.get("Subject Code", ""),
                        "co_vals": co_vals, "total": total,
                        "final_total": d.get("final_total", total),
                        "threshold": d.get("threshold", 0),
                        "max_marks": d.get("Max Marks", max_marks_input),
                    }

                normalised     = [n for n in [normalise(d) for d in all_data] if str(n["name"]).strip() and str(n["name"]).strip().lower() not in ("nan","none","")]
                max_co         = max((len(n["co_vals"]) for n in normalised), default=0)
                co_headers     = [f"CO{i+1}" for i in range(max_co)]
                formatted_data = []

                for n in normalised:
                    co_vals   = n["co_vals"]
                    obtained  = n["total"]
                    max_marks = n["max_marks"]
                    att_pct   = min((obtained / max_marks * 100), 100) if max_marks > 0 else 0
                    att_level = 3 if att_pct>=70 else (2 if att_pct>=60 else (1 if att_pct>=50 else 0))
                    row = {
                        "Student Name": n["name"],      "Reg No": n["reg_no"],
                        "Subject Name": n["subject"],   "Subject Code": n["code"],
                        "Attendance":   n["status"],    "Max Marks": max_marks,
                        "Obtained Marks": obtained,     "Final Marks": n["final_total"],
                        "Threshold":    n["threshold"], "Attainment %": round(att_pct, 2),
                        "Attainment Level": att_level,
                    }
                    for i in range(max_co):
                        row[f"CO{i+1}"] = co_vals[i] if i < len(co_vals) else 0
                    formatted_data.append(row)

                file_name = f"{st.session_state.teacher}_students_data.xlsx"
                wb_dl = Workbook(); ws_dl = wb_dl.active; ws_dl.title = "Student Data"
                headers = ["Student Name","Reg No","Subject Name","Subject Code","Attendance",
                           "Max Marks","Obtained Marks","Final Marks","Threshold",
                           "Attainment %","Attainment Level"] + co_headers
                for col_num, header in enumerate(headers, 1):
                    cell = ws_dl.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")
                for row_num, row_data in enumerate(formatted_data, start=2):
                    for col_num, key in enumerate(headers, 1):
                        ws_dl.cell(row=row_num, column=col_num, value=row_data.get(key, ""))
                for col in ws_dl.columns:
                    max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                    ws_dl.column_dimensions[get_column_letter(col[0].column)].width = max_length + 3
                dl_buf = io.BytesIO(); wb_dl.save(dl_buf); dl_buf.seek(0)
                st.download_button("⬇️ Download Excel File", dl_buf, file_name=file_name,
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        if calculate_clicked:
            # Use only THIS session's data
            session_manual = st.session_state.get("session_manual_students", [])
            session_upload_df = st.session_state.get("session_upload_df", None)

            session_uploads = []
            if session_upload_df is not None:
                clean_df2 = session_upload_df[
                    session_upload_df.apply(
                        lambda r: str(r.get("Student Name", r.get("student_name", r.get("Name", "")))).strip().lower()
                                  not in ("", "nan", "none"),
                        axis=1
                    )
                ]
                for row in clean_df2.to_dict(orient="records"):
                    row["teacher"] = st.session_state.get("teacher", "")
                    row["status"]  = str(row.get("Attendance", "Present")).strip()
                    session_uploads.append(row)

            if len(session_manual) == 0 and len(session_uploads) == 0:
                st.warning("⚠️ No data for this session. Upload a file or add students manually first.")
                calculate_clicked = False

            # Temporarily pass session data to calculate_attainment via a flag
            st.session_state["_session_students"] = session_manual
            st.session_state["_session_uploads"]  = session_uploads

            threshold_val = int(threshold_global)
            df, summary, co_excel_buffer = calculate_attainment(
                threshold_val, max_marks_input,
                co_maxes=co_maxes_input,
                teacher=st.session_state.teacher,
                session_students=st.session_state.get("_session_students", []),
                session_uploads=st.session_state.get("_session_uploads", []),
                program=program_name,       semester=semester_val,
                course_code=course_code_val, course_name=course_name_val,
                department=department_name,  university=university_name,
                exam_label=exam_label_val
            )
            st.session_state["co_excel_buffer"] = co_excel_buffer

            if df.empty:
                st.warning("⚠️ No valid student data found. Check your uploaded file.")
                calculate_clicked = False
            else:
                # df already has Total % from calculate.py; recompute Status with UI threshold
                if "Total %" not in df.columns:
                    df["Total %"] = df["Total Marks"].apply(
                        lambda m: round(m / max_marks_input * 100, 2) if max_marks_input > 0 else 0
                    )
                df["Status"] = df.apply(
                    lambda row: "Absent" if row["Status"] == "Absent"
                                else ("Pass" if row["Total %"] >= threshold_val else "Fail"), axis=1
                )

                passed = (df["Status"] == "Pass").sum()
                failed = (df["Status"] == "Fail").sum()
                absent = (df["Status"] == "Absent").sum()
                total_present  = passed + failed
                attainment_pct = round(passed / total_present * 100, 2) if total_present > 0 else 0

                st.subheader("📊 Attainment Overview", anchor=False)
                col_c1, col_c2, col_c3, col_c4 = st.columns(4)
                col_c1.metric("✅ Passed",     passed)
                col_c2.metric("❌ Failed",     failed)
                col_c3.metric("🚫 Absent",     absent)
                col_c4.metric("🎯 Attainment", f"{attainment_pct}%")

                present_df = df[df["Status"] != "Absent"].copy()
                present_df = present_df.sort_values("Total %", ascending=False).reset_index(drop=True)
                scores = present_df["Total %"].tolist()
                avg    = round(sum(scores) / len(scores), 1) if scores else 0

                fig, axes = plt.subplots(1, 2, figsize=(13, 5))
                fig.patch.set_facecolor("#1e293b")

                ax1 = axes[0]; ax1.set_facecolor("#1e293b")
                bars = ax1.bar(["Passed","Failed","Absent"],[passed,failed,absent],
                               color=["#22c55e","#ef4444","#94a3b8"],width=0.5,edgecolor="none")
                for bar, val in zip(bars,[passed,failed,absent]):
                    ax1.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.3,
                             str(int(val)),ha="center",va="bottom",color="white",fontsize=13,fontweight="bold")
                ax1.set_title("Pass / Fail / Absent",color="white",fontsize=12,fontweight="bold",pad=12)
                ax1.tick_params(colors="white",labelsize=11)
                ax1.set_ylim(0,max(passed,failed,absent,1)*1.25)
                for spine in ax1.spines.values(): spine.set_visible(False)
                ax1.yaxis.set_visible(False); ax1.grid(False)

                ax2 = axes[1]; ax2.set_facecolor("#1e293b")
                bar_colors = ["#22c55e" if s>=threshold_val else "#ef4444" for s in scores]
                ax2.bar(range(len(scores)),scores,color=bar_colors,width=1.0,edgecolor="none",alpha=0.85)
                ax2.axhline(y=threshold_val,color="#facc15",linewidth=2,linestyle="--",zorder=5)
                ax2.axhline(y=avg,color="#60a5fa",linewidth=2,linestyle="-",zorder=5)
                ax2.set_title("Score Distribution (High to Low)",color="white",fontsize=12,fontweight="bold",pad=12)
                ax2.set_xlabel("Students ranked by score",color="#94a3b8",fontsize=10)
                ax2.set_ylabel("Score %",color="#94a3b8",fontsize=10)
                ax2.set_ylim(0,110); ax2.set_xticks([])
                ax2.tick_params(colors="#94a3b8",labelsize=9)
                for spine in ax2.spines.values(): spine.set_visible(False)
                ax2.grid(axis="y",color="#374151",linewidth=0.6,alpha=0.5)

                import matplotlib.patches as mpatches
                ax2.legend(handles=[
                    mpatches.Patch(color="#22c55e", label=f"Above threshold ({passed})"),
                    mpatches.Patch(color="#ef4444", label=f"Below threshold ({failed})"),
                    plt.Line2D([0],[0],color="#facc15",linewidth=2,linestyle="--",label=f"Threshold: {threshold_val}%"),
                    plt.Line2D([0],[0],color="#60a5fa",linewidth=2,label=f"Class Avg: {avg}%")],
                    facecolor="#0f172a",edgecolor="#334155",labelcolor="white",fontsize=8.5,loc="upper right")

                plt.tight_layout(pad=2)
                st.pyplot(fig)

                co_buf = st.session_state.get("co_excel_buffer")
                if co_buf:
                    st.download_button(
                        "⬇️ Download CO Attainment Report",
                        co_buf, file_name="CO_Attainment.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

    # ── LOGOUT ─────────────────────────────────────────
    st.markdown("---")
    if st.button("🚪 Logout", use_container_width=True):
        st.session_state.logged_in = False
        st.session_state.page = "home"
        st.rerun()

# ---------------- FOOTER ----------------
show_footer()
